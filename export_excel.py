import json
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

SRC = r'C:\Users\qingf\Desktop\shop\all_items.json'
OUT = r'C:\Users\qingf\Desktop\shop\szwego_export.xlsx'

print('Loading JSON...')
with open(SRC, 'r', encoding='utf-8') as f:
    items = json.load(f)

print(f'{len(items)} items')

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '商品数据'

# Headers
headers = [
    '序号', 'goods_id', '标题', '副标题', '发布日期', '图片数量',
    '图片链接(原图)', '图片链接(缩略图)',
    '详情页链接', '小程序路径', '店铺名', 'SKU', '状态'
]
for col, h in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=h)

# Column widths
widths = [6, 40, 60, 20, 14, 8, 80, 80, 50, 60, 20, 20, 8]
for col, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = w

# Freeze header
ws.freeze_panes = 'A2'

# Write data
row = 2
empty_count = 0
for idx, item in enumerate(items):
    imgs_src = item.get('imgsSrc') or []
    imgs = item.get('imgs') or []
    title = (item.get('title') or '').strip()
    sub_title = (item.get('subTitle') or '').strip()
    img_count = len(imgs_src)
    link = item.get('link', '')
    detail_url = f'https://www.szwego.com{link}' if link else ''
    time_str = item.get('time', '')
    sku = item.get('sku', {})
    sku_name = sku.get('name', '') if isinstance(sku, dict) else str(sku)

    # Skip completely empty items
    if not title and img_count == 0:
        empty_count += 1

    ws.cell(row=row, column=1, value=idx + 1)
    ws.cell(row=row, column=2, value=item.get('goods_id', ''))
    ws.cell(row=row, column=3, value=title[:500] if title else '')
    ws.cell(row=row, column=4, value=sub_title[:200] if sub_title else '')
    ws.cell(row=row, column=5, value=time_str)
    ws.cell(row=row, column=6, value=img_count)
    ws.cell(row=row, column=7, value='\n'.join(imgs_src[:20]) if imgs_src else '')
    ws.cell(row=row, column=8, value='\n'.join(imgs[:20]) if imgs else '')
    ws.cell(row=row, column=9, value=detail_url)
    ws.cell(row=row, column=10, value=item.get('miniapp_path', ''))
    ws.cell(row=row, column=11, value=item.get('shop_name', ''))
    ws.cell(row=row, column=12, value=sku_name)
    ws.cell(row=row, column=13, value=item.get('status', ''))

    row += 1
    if row % 5000 == 0:
        print(f'  Written {row - 2} rows...')

wb.save(OUT)
wb.close()

print(f'\nExported: {row - 2} rows (skipped {empty_count} empty items)')
print(f'Saved to: {OUT}')
