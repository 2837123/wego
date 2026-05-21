import openpyxl, json, sys
from openpyxl.utils import get_column_letter

if len(sys.argv) < 3:
    print('Usage: python export_template.py <data.json> <output.xlsx>')
    sys.exit(1)

data_file = sys.argv[1]
out_file = sys.argv[2]
import os
# Look for template next to script
_script_dir = os.path.dirname(os.path.abspath(__file__))
_tpl_local = os.path.join(_script_dir, 'shitimoban.xlsx')
if os.path.exists(_tpl_local):
    tpl_file = _tpl_local
else:
    print('Error: Template file not found at "' + _tpl_local + '". Please copy shitimoban.xlsx to the project directory.')
    sys.exit(1)

with open(data_file, 'r', encoding='utf-8') as f:
    export_data = json.load(f)

products = export_data.get('products', [])

# Copy template headers and widths
wb_tpl = openpyxl.load_workbook(tpl_file)
ws_tpl = wb_tpl['Sheet1']
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Sheet1'

for col in range(1, 32):
    ws.cell(row=1, column=col, value=ws_tpl.cell(row=1, column=col).value)
    w = ws_tpl.column_dimensions[get_column_letter(col)].width
    if w:
        ws.column_dimensions[get_column_letter(col)].width = w
wb_tpl.close()

# ---- Hardcoded defaults (from template examples) ----
HARD = {
    'support':    '承诺包退;厂家直发;假一赔四',  # Col 10 商品支持
    'stock':      1000,                    # Col 18 库存
    'weight':     0.3,                     # Col 19 重量(千克)
    'volume':     0,                       # Col 20 体积
    'vip':        '会员价;会员等级折扣;会员卡折扣;会员卡价',  # Col 21 会员权益
    'show_stock': '隐藏',                  # Col 22 显示库存
    'show_sales': '显示',                  # Col 23 显示销量
    'logistics':  '快递发货;同城配送;上门自提',  # Col 24 物流支持
    'ship_method':'快递发货',              # Col 25 下单默认配送方式
    'freight':    '运费模板:2234',          # Col 26 快递运费
    'status':     '上架售卖',              # Col 28 商品状态
    'poster':     '',                      # Col 29 商品海报
    'form':       '无',                    # Col 30 商品表单
    'channel':    '公众号;小程序;PC;抖音;H5', # Col 31 出售渠道
    'price_rate': 2,                       # 价格倍率
}

row = 2
for i, pr in enumerate(products):
    ai = (pr.get('aiFields') or {})
    imgs = pr.get('imgs', []) or []
    img_str = ''.join(['[' + u + ']' for u in imgs[:5]])       # 商品图片只放前5张
    img_str_full = ''.join(['[' + u + ']' for u in imgs])    # 详情放全部
    first_img = imgs[0] if imgs else ''
    copy = (pr.get('copy', '') or '').replace('\n', ' ').strip()
    style = pr.get('style', '') or ''
    raw_price = pr.get('price', '') or ''
    try:
        price = float(raw_price) * HARD['price_rate']
        # 划线价 = 原始价格 * 2.3，四舍五入取整，9结尾
        raw_p = float(raw_price)
        line_p = round(raw_p * 2.3)
        line_price = round(line_p / 10) * 10 - 1
    except:
        price = ''
        line_price = ''

    # ---- Col 1: 商品名称 ----
    title = (ai.get('title') or '').strip()
    if not title and ai.get('title') is not None:
        title = ''  # explicitly empty if AI returned empty (disabled)
    if not title and ai.get('title') is None:
        title = copy.split('\n')[0].strip()[:200] if copy else ('商品 #' + style)

    # ---- Col 2: 副标题 ----
    subtitle = (ai.get('subtitle') or '').strip()[:50]

    # ---- Col 3: 短标题 ----
    short = (ai.get('short') or '').strip()[:25]

    # ---- Col 4: 标题标签 ----
    tag = ai.get('tag')
    if tag is None:
        tag = '人气爆款'  # not processed by AI, use default
    else:
        tag = tag.strip()

    # ---- Col 5: 商品图片 ----
    # Already in [url][url] format

    # ---- Col 6: 搜索关键词 ----
    keywords = ai.get('keywords')
    if keywords is None or not keywords.strip():
        # Fallback: derive from product category
        cate_name = (pr.get('category') or '').replace('[', '').replace(']', '').split(';')[-1] if pr.get('category') else ''
        kw_map = {
            '牛仔裤':'牛仔裤;裤子;女装','休闲裤':'休闲裤;裤子;女装','短裤':'短裤;女装',
            '连衣裙':'连衣裙;裙子;女装','短裙':'短裙;裙子;女装','长裙':'长裙;裙子;女装',
            'T恤':'T恤;短袖;女装','衬衫':'衬衫;上衣;女装','卫衣':'卫衣;上衣;女装',
            '外套':'外套;开衫;女装','大衣':'大衣;外套;女装','针织衫':'针织衫;毛衣;女装',
            '背心':'背心;吊带;女装','打底':'打底;上衣;女装'
        }
        keywords = kw_map.get(cate_name, '女装;新款')
    else:
        keywords = keywords.strip()

    # ---- Col 7: 商品特征标签 ----
    features = ai.get('features')
    if features is None or not features.strip():
        features = '厂家直销'
    else:
        features = features.strip()

    # ---- Col 8: 商品分类 (一级分类，如 [牛仔裤]) ----
    category = pr.get('category', '')
    if not category:
        category = ai.get('category')
        if category is None or not category.strip():
            category = ''
    # Extract just the subcategory name (remove [女装; prefix)
    if category:
        cat = category.replace('[', '').replace(']', '').strip()
        if ';' in cat:
            cat = cat.split(';')[-1].strip()
        category = '[' + cat + ']'
    else:
        category = ''

    # ---- Hardcoded fields ----
    grouping = '2026'                     # Col 9 商品分组
    support = HARD['support']             # Col 10
    code = style                          # Col 12
    barcode = ''                          # Col 13
    spec_img = first_img                  # Col 14
    sell_price = price                    # Col 15
    # line_price computed above from raw price * 2.3
    cost_price = ''                       # Col 17
    stock = HARD['stock']                 # Col 18
    weight = HARD['weight']               # Col 19
    volume = HARD['volume']               # Col 20
    vip = HARD['vip']                     # Col 21
    show_stock = HARD['show_stock']       # Col 22
    show_sales = HARD['show_sales']       # Col 23
    logistics = HARD['logistics']         # Col 24
    ship_method = HARD['ship_method']     # Col 25
    freight = HARD['freight']             # Col 26
    status = HARD['status']               # Col 28
    poster = HARD['poster']               # Col 29
    form = HARD['form']                   # Col 30
    channel = HARD['channel']             # Col 31

    # ---- Col 27: 移动端商品详情 (auto-generated, no AI) ----
    detail = pr.get('detail', '') or (img_str_full + '\n' + copy.replace('\n', '<br>'))
    detail = detail[:5000]

    # Write 3 rows per product (S / M / L)
    for size in ['S', 'M', 'L']:
        spec = '[颜色:图片色][尺码:' + size + ']'

        ws.cell(row=row, column=1, value=title)
        ws.cell(row=row, column=2, value=subtitle)
        ws.cell(row=row, column=3, value=short)
        ws.cell(row=row, column=4, value=tag)
        ws.cell(row=row, column=5, value=img_str)
        ws.cell(row=row, column=6, value=keywords)
        ws.cell(row=row, column=7, value=features)
        ws.cell(row=row, column=8, value=category)
        ws.cell(row=row, column=9, value=grouping)
        ws.cell(row=row, column=10, value=support)
        ws.cell(row=row, column=11, value=spec)
        ws.cell(row=row, column=12, value=code)
        ws.cell(row=row, column=13, value=barcode)
        ws.cell(row=row, column=14, value=spec_img)
        ws.cell(row=row, column=15, value=sell_price)
        ws.cell(row=row, column=16, value=line_price)
        ws.cell(row=row, column=17, value=cost_price)
        ws.cell(row=row, column=18, value=stock)
        ws.cell(row=row, column=19, value=weight)
        ws.cell(row=row, column=20, value=volume)
        ws.cell(row=row, column=21, value=vip)
        ws.cell(row=row, column=22, value=show_stock)
        ws.cell(row=row, column=23, value=show_sales)
        ws.cell(row=row, column=24, value=logistics)
        ws.cell(row=row, column=25, value=ship_method)
        ws.cell(row=row, column=26, value=freight)
        ws.cell(row=row, column=27, value=detail)
        ws.cell(row=row, column=28, value=status)
        ws.cell(row=row, column=29, value=poster)
        ws.cell(row=row, column=30, value=form)
        ws.cell(row=row, column=31, value=channel)
        row += 1

wb.save(out_file)
wb.close()
print(f'OK:{row - 2}')
